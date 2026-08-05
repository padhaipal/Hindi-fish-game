#!/usr/bin/env python3
r"""
Add `caption` and `flags` columns to a per-photo sheet (from ocr_timestamps.py)
using the WhatsApp chat export -- WITHOUT re-running OCR (takes seconds).

For each photo it looks up the message that posted it and pulls:
  * caption -- the text the teacher wrote with the photo (e.g. "Class start photo",
    "Ishrat ki class ending", "Afreen ki class vizti").
  * flags   -- tolerant keyword detection over that caption, any of:
       start / end / visit / mentions:<Teacher>
    (spelling-tolerant: catches vizti/vist, stetion, ending/over, etc.)

These give a later AI/human pass clean signals to finish the judgement calls
(which photo is a class start vs end, and when the supervisor is just visiting
someone else's class) instead of re-reading 24k lines of chat.

Usage:
    python enrich_captions.py June_timestamps.xlsx --chat "chat.txt" -o June_timestamps_enriched.xlsx
"""

import argparse
import datetime
import re
import sys

from openpyxl import load_workbook

# Teachers we recognise in captions (caption name -> canonical). Mariyam = Ishrat.
TEACHER_NAMES = {
    "ishrat": "Ishrat", "mariyam": "Ishrat", "afreen": "Afreen",
    "anjum": "Anjum", "nahid": "Nahid", "heena": "Heena", "mantasha": "Mantasha",
}

HDR = re.compile(
    r'^(\d{1,2})/(\d{1,2})/(\d{2,4}),\s+(\d{1,2}):(\d{2})(?::\d{2})?\s*'
    r'([apAP][. ]?[mM])?\s*[-–]\s*(.*)$')
ATTACH = re.compile(r'([A-Za-z0-9._-]+\.\w+)\s*\(file attached\)')


def parse_chat_full(path):
    """Map {filename: (datetime, sender, caption)} including caption lines."""
    mapping = {}
    cur = None

    def flush(c):
        if not c or not c["fname"]:
            return
        try:
            dt = datetime.datetime(c["y"], c["mo"], c["d"], c["h"], c["mi"])
        except ValueError:
            return
        mapping.setdefault(c["fname"], (dt, c["sender"], " ".join(c["caption"]).strip()))

    with open(path, encoding="utf-8", errors="replace") as fh:
        for raw in fh:
            line = raw.rstrip("\n")
            m = HDR.match(line)
            if m:
                flush(cur)
                d, mo, y, h, mi, ap, body = m.groups()
                y = int(y); y = y + 2000 if y < 100 else y
                h = int(h)
                if ap:
                    ap = ap.lower().replace(".", "").replace(" ", "")
                    if ap == "pm" and h != 12:
                        h += 12
                    if ap == "am" and h == 12:
                        h = 0
                sender, content = (body.split(": ", 1) + [""])[:2] if ": " in body else ("", body)
                am = ATTACH.search(content)
                cap = []
                # any text after "(file attached)" on the same line is a caption too
                if am:
                    tail = content[am.end():].strip()
                    if tail:
                        cap.append(tail)
                cur = {"y": y, "mo": int(mo), "d": int(d), "h": h, "mi": int(mi),
                       "sender": sender.strip(), "fname": am.group(1) if am else None,
                       "caption": cap}
            elif cur is not None and cur["fname"] and line.strip():
                cur["caption"].append(line.strip())
        flush(cur)
    return mapping


def detect_flags(caption):
    """Return a semicolon-joined flag string from a caption."""
    c = (caption or "").lower()
    flags = []
    if re.search(r'\b(start|shuru|strt|statr)', c):
        flags.append("start")
    # \b avoids matching "end" inside "attendance"
    if re.search(r'\b(end|ending|over|khatam|finish|samapt)', c):
        flags.append("end")
    if re.search(r'\b(visit|vizit|vizti|vist|nirikshan)', c):
        flags.append("visit")
    mentioned = []
    for key, name in TEACHER_NAMES.items():
        if key in c and name not in mentioned:
            mentioned.append(name)
    if mentioned:
        flags.append("mentions:" + "/".join(mentioned))
    return "; ".join(flags)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("photos", help="Per-photo .xlsx from ocr_timestamps.py")
    ap.add_argument("--chat", required=True, help="WhatsApp chat export .txt")
    ap.add_argument("-o", "--output", required=True, help="Output .xlsx")
    args = ap.parse_args()

    chat = parse_chat_full(args.chat)
    caps = {fn: cap for fn, (dt, sender, cap) in chat.items()}

    wb = load_workbook(args.photos)
    ws = wb.active
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        fn_col = header.index("filename") + 1
    except ValueError:
        sys.exit("No 'filename' column found - is this the ocr_timestamps output?")

    # add columns (reuse if already present)
    def col_for(name):
        if name in header:
            return header.index(name) + 1
        ws.cell(row=1, column=ws.max_column + 1, value=name)
        header.append(name)
        return ws.max_column

    cap_col = col_for("caption")
    flag_col = col_for("flags")

    n_cap = n_flag = 0
    for r in range(2, ws.max_row + 1):
        fn = ws.cell(row=r, column=fn_col).value
        if not fn:
            continue
        cap = caps.get(str(fn), "")
        if cap:
            n_cap += 1
        flags = detect_flags(cap)
        if flags:
            n_flag += 1
        ws.cell(row=r, column=cap_col, value=cap)
        ws.cell(row=r, column=flag_col, value=flags)

    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  captions found : {n_cap}")
    print(f"  rows with flags: {n_flag}")


if __name__ == "__main__":
    main()
