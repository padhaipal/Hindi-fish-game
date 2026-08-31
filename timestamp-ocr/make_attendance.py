#!/usr/bin/env python3
r"""
Reshape the per-photo timestamp sheet (from ocr_timestamps.py, ideally after
enrich_captions.py) into the teacher-attendance grid, matching the layout of
teacher_attendance_format.xlsx.

The class structure is READ FROM THE TEMPLATE (row "Teacher" and row "Time"), so
the number of classes, their teachers, times and columns can change month to
month without editing this script -- just supply that month's template.

For each class on each day:
  * Start = first photo, End = last photo, Total = End - Start.
  * If caption `flags` are present (from enrich_captions.py), a `start`/`end`
    label sets the real Start/End.
  * Photos are assigned to a class by the template's scheduled time (nearest
    window). A teacher whose two classes share the SAME time (e.g. Anjum) is
    split by `start` flags: one start that day -> one class ran (the rest is that
    class's own end), so the second class stays blank.

Highlighting:
  * RED    = teacher absent (no photos) on a working day (Mon-Sat).
  * ORANGE = only one photo that day (a start or end, but not both).
  * GREEN  = the supervisor (Ishrat) visited this class that day.
  * A class with no photos and no visits all month is left blank.
A "Supervisions" row below the dates totals each class's visits for the month.

Usage:
    python make_attendance.py June_timestamps_enriched.xlsx \
        --template "teacher_attendance_format.xlsx" -o June_attendance.xlsx
"""

import argparse
import datetime
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

# ============================ CONFIG =========================================
SUPERVISOR = "Ishrat"            # whose "visit" photos supervise other classes
SENDER_RENAME = {"mariyam": "Ishrat"}   # chat sender -> teacher name
GAP_SPLIT_MIN = 75               # (unused fallback) minutes gap for same-time split
RED = PatternFill("solid", fgColor="FFFFC7CE")     # absent
ORANGE = PatternFill("solid", fgColor="FFFFC000")  # only one photo
GREEN = PatternFill("solid", fgColor="FFC6EFCE")   # supervised (visit)
NOFILL = PatternFill()

# Correct a class's time window when the template's scheduled time is stale.
# Keyed by a lowercase substring of the class's Teacher label; value is
# ("H:MM", "H:MM"). Ishrat's real classes are reversed vs the schedule: her
# "1st" (home) runs in the afternoon and her "2nd" (Daliganj) in the morning.
WINDOW_OVERRIDE = {
    "ishrat 1st": ("11:00", "20:00"),
    "ishrat 2nd": ("6:00", "11:00"),
}
# =============================================================================


def mins(t):
    m = re.match(r'(\d{1,2}):(\d{2})', t.strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def parse_window(text):
    if not text:
        return None
    parts = re.split(r'[-–]', str(text))
    if len(parts) != 2:
        return None
    a, b = mins(parts[0]), mins(parts[1])
    return (a, b) if a is not None and b is not None else None


def base_name(label):
    m = re.match(r'\s*([A-Za-z]+)', str(label or ""))
    return m.group(1).capitalize() if m else None


def read_template_classes(ws):
    """Return (classes, date_row). classes: list of dicts col/label/base/win in order."""
    labels = {}
    for r in range(1, 20):
        v = ws.cell(row=r, column=1).value
        if v:
            labels[str(v).strip().lower()] = r
    teacher_row = labels.get("teacher")
    time_row = labels.get("time")
    date_row = labels.get("date")
    if not (teacher_row and date_row):
        sys.exit("Template must have 'Teacher' and 'Date' labels in column A.")
    classes = []
    col = 2
    while col <= ws.max_column:
        label = ws.cell(row=teacher_row, column=col).value
        if label and str(label).strip():
            win = parse_window(ws.cell(row=time_row, column=col).value) if time_row else None
            low = str(label).strip().lower()
            for key, (a, b) in WINDOW_OVERRIDE.items():
                if key in low:
                    win = (mins(a), mins(b))
                    break
            classes.append({"col": col, "label": str(label).strip(),
                            "base": base_name(label), "win": win})
        col += 4
    if not classes:
        sys.exit("No class columns found in template (row 'Teacher' empty).")
    return classes, date_row


def make_canon(bases):
    def canon(sender):
        s = (sender or "").strip().lower()
        for k, v in SENDER_RENAME.items():
            if k in s:
                return v
        for b in bases:
            if b.lower() in s:
                return b
        return None
    return canon


def parse_dt(be):
    if isinstance(be, datetime.datetime):
        return be
    if isinstance(be, str) and be.strip():
        for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.datetime.strptime(be.strip(), f)
            except ValueError:
                pass
    return None


def load_photos(path, canon):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    H = {n: i for i, n in enumerate(hdr)}
    if "teacher" not in H:
        sys.exit("Input has no 'teacher' column - is this the ocr_timestamps output?")
    out = []
    for row in it:
        if not row or all(c is None for c in row):
            continue
        teacher = canon(row[H["teacher"]] if H["teacher"] < len(row) else "")
        if not teacher:
            continue
        dt = parse_dt(row[H["best_estimate"]] if H.get("best_estimate") is not None else None)
        if not dt:
            continue
        flags = ""
        if H.get("flags") is not None and H["flags"] < len(row) and row[H["flags"]]:
            flags = str(row[H["flags"]]).lower()
        out.append((teacher, dt, flags))
    return out


def assign(classes, recs):
    """recs: sorted [(minute, flags)]. Return {col: [(minute, flags)]}."""
    out = {c["col"]: [] for c in classes}
    if not recs:
        return out
    if len(classes) == 1:
        out[classes[0]["col"]] = recs[:]
        return out
    wins = [c["win"] for c in classes]
    if len(set(wins)) == 1:  # same-time group (e.g. Anjum): split by start flags
        starts = [i for i, (m, f) in enumerate(recs) if "start" in f]
        nsess = min(max(1, len(starts)), len(classes))
        if nsess <= 1:
            out[classes[0]["col"]] = recs[:]
        else:
            idxs = starts[:nsess]
            segs = []
            for si in range(len(idxs)):
                a = idxs[si]
                b = idxs[si + 1] if si + 1 < len(idxs) else len(recs)
                segs.append(recs[a:b])
            if idxs[0] > 0:
                segs[0] = recs[:idxs[0]] + segs[0]
            for ci, seg in enumerate(segs[:len(classes)]):
                out[classes[ci]["col"]] = seg
        return out

    def dist(t, c):
        if not c["win"]:
            return 10 ** 9
        s, e = c["win"]
        return 0 if s <= t <= e else (s - t if t < s else t - e)
    for m, f in recs:
        out[min(classes, key=lambda c: dist(m, c))["col"]].append((m, f))
    return out


def hhmm(m):
    return f"{m // 60}:{m % 60:02d}"


def start_end(recs):
    times = [m for m, _ in recs]
    starts = [m for m, f in recs if "start" in f]
    ends = [m for m, f in recs if "end" in f]
    s = min(starts) if starts else times[0]
    e = max(ends) if ends else times[-1]
    if e < s:
        s, e = times[0], times[-1]
    return s, e


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("photos")
    ap.add_argument("--template", required=True)
    ap.add_argument("-o", "--output", default="attendance.xlsx")
    args = ap.parse_args()

    wb = load_workbook(args.template)
    ws = wb.active
    classes, date_row = read_template_classes(ws)
    bases = {c["base"] for c in classes if c["base"]}
    canon = make_canon(bases)
    by_base = {}
    for c in classes:
        by_base.setdefault(c["base"], []).append(c)

    photos = load_photos(args.photos, canon)
    if not photos:
        sys.exit("No photos with a recognized teacher were found.")

    # Split visits (supervisor visiting another class) from teaching photos.
    by_key = {}          # (base, date) -> [(minute, flags)]
    visits = []          # (mentioned_base, date, minute) -- teacher named in caption
    ctx_visits = []      # (date, minute) -- no teacher named; infer from context
    all_dates = set()
    n_visit = 0
    for teacher, dt, flags in photos:
        d = dt.date()
        all_dates.add(d)
        if "visit" in flags:
            n_visit += 1
            mentioned = re.search(r'mentions:([a-z/]+)', flags)
            names = [n.capitalize() for n in mentioned.group(1).split("/")] if mentioned else []
            targets = [n for n in names if n in bases and n != teacher]
            if targets:
                for n in targets:
                    visits.append((n, d, dt.hour * 60 + dt.minute))
            else:
                ctx_visits.append((d, dt.hour * 60 + dt.minute))
            continue
        by_key.setdefault((teacher, d), []).append((dt.hour * 60 + dt.minute, flags))

    # Assign teaching photos to class columns.
    cell_recs = {}
    teaching_cols = set()
    for (base, d), recs in by_key.items():
        if base not in by_base:
            continue
        recs.sort()
        for col, cr in assign(by_base[base], recs).items():
            if cr:
                cell_recs[(col, d)] = sorted(cr)
                teaching_cols.add(col)

    # Attribute each visit to the mentioned teacher's class column for that date.
    supervised = {}      # (col, date) -> count
    superv_total = {}    # col -> count
    n_ctx_resolved = 0
    ish_cols = {c["col"] for c in classes if c["base"] == SUPERVISOR}

    def record(col, d):
        supervised[(col, d)] = supervised.get((col, d), 0) + 1
        superv_total[col] = superv_total.get(col, 0) + 1

    for base, d, minute in visits:
        cls = by_base.get(base, [])
        if not cls:
            continue
        def dist(c):
            if not c["win"]:
                return 10 ** 9
            s, e = c["win"]
            return 0 if s <= minute <= e else (s - minute if minute < s else minute - e)
        chosen = min(cls, key=lambda c: (0 if c["col"] in teaching_cols else 1, dist(c)))
        record(chosen["col"], d)

    # Unnamed visits: infer the class from context -- the non-Ishrat class whose
    # session that day was running closest to the visit time.
    for d, minute in ctx_visits:
        cands = []
        for (col, dd), recs in cell_recs.items():
            if dd != d or col in ish_cols:
                continue
            t0 = min(m for m, _ in recs)
            t1 = max(m for m, _ in recs)
            gap = 0 if t0 <= minute <= t1 else min(abs(minute - t0), abs(minute - t1))
            cands.append((gap, col))
        if cands:
            cands.sort()
            record(cands[0][1], d)
            n_ctx_resolved += 1

    active_cols = teaching_cols | set(superv_total)
    lo, hi = min(all_dates), max(all_dates)
    dates = [lo + datetime.timedelta(days=i) for i in range((hi - lo).days + 1)]

    # Clear old data rows -- but only the Date column and the class columns
    # (Start/End/Total/Prize). Leave any other columns (e.g. "Daily Totals" /
    # "Weekly" summary columns the user maintains) untouched.
    clear_cols = {1}
    for c in classes:
        clear_cols.update(range(c["col"], c["col"] + 4))
    for r in range(date_row + 1, ws.max_row + 1):
        for c in clear_cols:
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = NOFILL

    center = Alignment(horizontal="center")
    for i, d in enumerate(dates):
        r = date_row + 1 + i
        ws.cell(row=r, column=1,
                value=f"{d.strftime('%a')}, {d.strftime('%b')} {d.day}, {d.strftime('%y')}")
        working = d.weekday() <= 5
        for c in classes:
            col = c["col"]
            if col not in active_cols:
                continue
            sc, ec, tc = (ws.cell(row=r, column=col), ws.cell(row=r, column=col + 1),
                          ws.cell(row=r, column=col + 2))
            for cell in (sc, ec, tc):
                cell.alignment = center
            recs = cell_recs.get((col, d))
            if recs:
                if len(recs) == 1:
                    sc.value = hhmm(recs[0][0])
                    fill = ORANGE
                else:
                    s, e = start_end(recs)
                    sc.value, ec.value, tc.value = hhmm(s), hhmm(e), hhmm(e - s)
                    fill = None
            else:
                fill = RED if working else None
            if (col, d) in supervised:      # supervision overrides absent/normal
                fill = GREEN
            if fill:
                for cell in (sc, ec, tc):
                    cell.fill = fill

    # Supervisions total row.
    tr = date_row + 1 + len(dates) + 1
    ws.cell(row=tr, column=1, value="Supervisions (visits by Ishrat)").font = Font(bold=True)
    for c in classes:
        if superv_total.get(c["col"]):
            cell = ws.cell(row=tr, column=c["col"], value=int(superv_total[c["col"]]))
            cell.number_format = "General"   # template cells are time-formatted; force plain count
            cell.fill = GREEN
            cell.alignment = center

    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  classes read from template: {len(classes)}  "
          f"({', '.join(c['label'][:14] for c in classes)})")
    print(f"  dates: {lo} .. {hi}  ({len(dates)} rows)")
    print(f"  active class-columns: {len(active_cols)} of {len(classes)}")
    print(f"  visits: {n_visit} ({len(visits)} named, {n_ctx_resolved} inferred by context, "
          f"{len(ctx_visits) - n_ctx_resolved} unresolved)")
    if superv_total:
        by_label = {c["col"]: c["label"] for c in classes}
        print("  supervisions per class: "
              + ", ".join(f"{by_label[col]}={n}" for col, n in superv_total.items()))


if __name__ == "__main__":
    main()
